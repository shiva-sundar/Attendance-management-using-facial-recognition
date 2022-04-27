
from tkinter import *
from tkinter.ttk import *
from PIL import Image,ImageTk
from face import capture_image
from tkinter import StringVar
import openpyxl
import os
import face
import cv2
from face import extract_faces
import face_recognition
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure



xl_path="database\\total.xlsx"

alphabets=['A','B','C','D','E','F']
cols=['rollno','name','img_path','total_present','total_absent','total_classes']


wb = openpyxl.load_workbook(xl_path)
sheet = wb.active 




   

def addStudent():
    add_student_window=Tk()
    add_student_window.title("Add Student")
    add_student_window.geometry("600x420")
    
    student_name_var=StringVar(add_student_window)
    image_path_var=StringVar(add_student_window)
    enrollment_number_var=StringVar(add_student_window)
    
    student_name_label = Label(add_student_window,text = "student name").place(x = 100,y = 100)
    student_name_entry = Entry(add_student_window,textvariable = student_name_var,width=30)
     
    enrollment_number_label= Label(add_student_window,text = "Enrollment Number").place(x = 100,y = 130)
    enrollment_number_entry = Entry(add_student_window,textvariable = enrollment_number_var,width=30 )
    
    def upload():
        
        image = face_recognition.load_image_file(r"{}".format(capture_image()))   #r"C:\Users\bsskt\OneDrive\Desktop\projects\Attendance\test3.jpg"
        face_locations = face_recognition.face_locations(image)
        print(face_locations)
        
        extracted_faces_image_names=set()
        
        image_name=str(enrollment_number_var.get())
        
        for i, face_location in enumerate(face_locations):
            # Print the location of each face in this image
            top, right, bottom, left = face_location
            
            # We can extract face using these 4 points
            face_image = image[top:bottom, left:right]
            pil_image = Image.fromarray(face_image)
            
            extracted_faces_image_names.add(image_name)
            saved_images="image_database\\" +image_name+".jpg"
            pil_image.save(saved_images)
        
        total_extracted_images = len(extracted_faces_image_names)
        Label(add_student_window,text = " "*100).place(x = 400,y = 195)
        
        if(total_extracted_images<1):
            Label(add_student_window,text = "Upload again!").place(x = 400,y = 195)
        else:
            Label(add_student_window,text = "Upload successfull!").place(x = 400,y = 195)
            
        print(total_extracted_images)
        return total_extracted_images
         
    upload_button=Button(add_student_window, text="Upload", command=upload)
    upload_button.place(x=300,y=190)
    
    
    
    def submit():
        student_name=student_name_var.get()
        enrollment_number=enrollment_number_var.get()
        image_path= "image_database\\" +enrollment_number+".jpg"
        
        Label(add_student_window,text = " "*100).place(x = 220,y = 280)
        print(student_name,image_path,enrollment_number)
        rows = sheet.max_row
        columns = sheet.max_column
        cell=(rows)+1
        #cell_obj = sheet_obj.cell(row = 1, column = 1)
        #sheet['A2']=100
        sheet['A'+str(cell)].value= enrollment_number
        sheet['B'+str(cell)].value=student_name
        sheet['C'+str(cell)].value=image_path
        sheet['D'+str(cell)].value=0
        sheet['E'+str(cell)].value=0
        sheet['F'+str(cell)].value=0
        wb.save(xl_path)
        
        Label(add_student_window,text = "Submitted successfully!").place(x = 220,y = 280)
        
        

      
    Submit = Button(add_student_window, text="Submit", command=submit)
    student_name_entry.place(x = 250,y = 100)
    enrollment_number_entry.place(x = 250, y = 130)
    Submit.place(x=300,y=230)
    add_student_window.mainloop()
    
    



def take_attendance():
    total_faces=extract_faces()
    faces_directory = 'extracted_faces'
    db_directory='image_database'
    
    cur_faces=0
    mark=set()

    for img in os.listdir(faces_directory):
        if cur_faces >= total_faces:
            break
        
        cur_faces+=1
        for img2 in os.listdir(db_directory):
            try:
         
                known_image = face_recognition.load_image_file(db_directory+'\\'+img2)
                unknown_image = face_recognition.load_image_file(faces_directory+'\\'+img)
                
                known_encoding = face_recognition.face_encodings(known_image)[0]
                unknown_encoding = face_recognition.face_encodings(unknown_image)[0]
                
                results = face_recognition.compare_faces([known_encoding], unknown_encoding)
                print(results[0])
                if results[0]:
                    roll=img2[:-4]
                    mark.add(roll)
                   
            except:
                continue
    
    print(mark)
    for i in range(2,sheet.max_row +1):
        print(sheet['F'+str(i)].value,type(sheet['F'+str(i)].value))
        sheet['F'+str(i)].value= int(sheet['F'+str(i)].value)+1
        for j in mark:
            if sheet['A'+str(i)].value == j:
                sheet['D'+str(i)].value = int(sheet['D'+str(i)].value)+1
                
    for i in range(2,sheet.max_row +1):
        sheet['E'+str(i)].value = int(sheet['F'+str(i)].value) - int(sheet['D'+str(i)].value)
         
    wb.save(xl_path)
        

        
'''
def reset():
    Label(show_attendance_window,text = " "*50).place(x = 420,y = 380)
    Label(show_attendance_window,text = " "*50).place(x = 540,y = 380)
    Label(show_attendance_window,text = " "*50).place(x = 660,y = 380)
    
    Label(show_attendance_window,text = " "*50).place(x = 450,y = 410)
    Label(show_attendance_window,text = " "*50).place(x = 570,y = 410)
    Label(show_attendance_window,text = " "*50).place(x = 690,y = 410)
    
reset()'''
    


def show_attendance():
    show_attendance_window=Tk()
    show_attendance_window.title("View attendance")
    show_attendance_window.geometry("900x500")
    
    enrollment_number_var=StringVar(show_attendance_window)
    
    enrollment_number_label= Label(show_attendance_window,text = "Enrollment Number").place(x = 10,y = 70)
    enrollment_number_entry = Entry(show_attendance_window,textvariable = enrollment_number_var,width=30 )
    enrollment_number_entry.place(x = 160, y = 70)
    
    def submit():
        enrollment_number=enrollment_number_var.get()
        flag=False 
        row_no=1
        valid_label= Label(show_attendance_window,text = " "*50).place(x = 190,y = 150)
        
        for i in range(2,sheet.max_row +1):
            print(sheet['A'+str(i)].value,enrollment_number)
            if enrollment_number == sheet['A'+str(i)].value:
                row_no=i
                flag=True
                break
            
        if (flag):
            frameChartsLT = Frame(show_attendance_window)
            frameChartsLT.pack(side=RIGHT, fill=BOTH)
    
            labels_pie = ['Classes Present','Classes Absent']
            values = [sheet['D'+str(row_no)].value,sheet['E'+str(row_no)].value]
            fig = Figure( dpi=90) # create a figure
            ax = fig.add_subplot(111) # add an Axes to the figure
            ax.pie(values, radius=1, labels=labels_pie,autopct='%0.2f%%',)
            chart1 = FigureCanvasTkAgg(fig,frameChartsLT)
            chart1.get_tk_widget().pack()
            
            Label(show_attendance_window,text = "Total classes").place(x = 420,y = 380)
            Label(show_attendance_window,text = "Total Present").place(x = 540,y = 380)
            Label(show_attendance_window,text = "Total Absent").place(x = 660,y = 380)
            Label(show_attendance_window,text = str(sheet['F'+str(row_no)].value)).place(x = 450,y = 410)
            Label(show_attendance_window,text = str(sheet['D'+str(row_no)].value)).place(x = 570,y = 410)
            Label(show_attendance_window,text = str(sheet['E'+str(row_no)].value)).place(x = 690,y = 410)
            
        else:
            valid_label= Label(show_attendance_window,text = "Enter valid Roll no !").place(x = 190,y = 150)
            
    Submit = Button(show_attendance_window, text="Submit", command=submit)
    Submit.place(x=200,y=100)
    
    show_attendance_window.mainloop()
    


main_window=Tk()
main_window.title("Attendance management System")
main_window.geometry("1100x620")

#exit
img = Image.open(r"icons\exit.png")
img = img.resize((75,60), Image.ANTIALIAS)
exit_icon =  ImageTk.PhotoImage(img)
Button(main_window, text = 'exit', image = exit_icon,command=main_window.destroy).place(x=980,y=40)
Label(main_window,text = "Exit").place(x = 1010,y = 110)

#take_attendance
img = Image.open(r"icons\take_attendance2.png")
img = img.resize((90,80), Image.ANTIALIAS)
take_attendance_icon =  ImageTk.PhotoImage(img)
Button(main_window, text = 'take_attendance', image = take_attendance_icon,command=take_attendance).place(x=100,y=270)
Label(main_window,text = "Take Attendance").place(x = 105,y = 360)

#addstudent

img = Image.open(r"icons\add_student.png")
img = img.resize((90,80), Image.ANTIALIAS)
add_student_icon =  ImageTk.PhotoImage(img)
Button(main_window, text = 'add_student', image = add_student_icon,command=addStudent).place(x=500,y=270)
Label(main_window,text = "Add Student").place(x = 510,y = 360)

#viewattendance
img = Image.open(r"icons\show_attendance.png")
img = img.resize((90,80), Image.ANTIALIAS)
show_attendance_icon =  ImageTk.PhotoImage(img)
Button(main_window, text = 'show_attendance', image = show_attendance_icon,command=show_attendance).place(x=900,y=270)
Label(main_window,text = "View Attendance").place(x = 905,y = 360)



main_window.mainloop()
